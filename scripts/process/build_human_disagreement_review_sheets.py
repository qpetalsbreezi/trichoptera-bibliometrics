#!/usr/bin/env python3
"""Build categorized disagreement review sheets (Excel + CSV) for human vs LLM eval."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

_SCRIPTS_DIR = Path(__file__).resolve().parent.parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

from lib.llm_validation import abstract_snippet, validation_dir  # noqa: E402
from lib.pipeline import paper_taxon_label  # noqa: E402

DEFAULT_JOINED = validation_dir() / "human_eval" / "joined_labeled.csv"
DEFAULT_HUMAN = (
    validation_dir() / "human_blind_coding" / "human_labeled_1376_20260830.csv"
)
DEFAULT_OUT = validation_dir() / "human_eval" / "disagreements_for_review"

FOCUS_KEEP = {"Primary focus", "Secondary mention"}


def gate_label(focus: str) -> str:
    return "keep" if focus in FOCUS_KEEP else "drop"


def enrich(joined: pd.DataFrame, human: pd.DataFrame) -> pd.DataFrame:
    df = joined.merge(
        human[["coding_row_id", "article_abstract"]],
        on="coding_row_id",
        how="left",
    )
    df["taxon_display"] = df["query_id"].map(paper_taxon_label)
    df["abstract_snippet"] = df["article_abstract"].map(lambda x: abstract_snippet(x, 500))
    for model in ("GPT", "GEM"):
        df[f"gate_{model}"] = df[f"focus_{model}"].map(gate_label)
    df["gate_H"] = df["focus_H"].map(gate_label)
    return df


def review_row(df: pd.DataFrame, model: str) -> pd.DataFrame:
    """One row per paper with human vs model columns for review."""
    m = model
    out = pd.DataFrame(
        {
            "coding_row_id": df["coding_row_id"],
            "taxon": df["taxon_display"],
            "year": df["Year"],
            "year_band": df["year_band"],
            "abstract_available": df["abstract_available"],
            "doi": df["DOI"],
            "title": df["Title"],
            "abstract_snippet": df["abstract_snippet"],
            "human_focus": df["focus_H"],
            "model_focus": df[f"focus_{m}"],
            "human_theme": df["theme_H"],
            "model_theme": df[f"theme_{m}"],
            "human_gate": df["gate_H"],
            "model_gate": df[f"gate_{m}"],
            "focus_pattern": df["focus_H"] + " → " + df[f"focus_{m}"],
            "theme_pattern": df["theme_H"] + " → " + df[f"theme_{m}"],
            "gate_pattern": df["gate_H"] + " → " + df[f"gate_{m}"],
            "reviewer_notes": "",
        }
    )
    return out


def add_flags(df: pd.DataFrame, model: str) -> pd.DataFrame:
    m = model
    out = df.copy()
    out["disagree_gate"] = out["gate_H"] != out[f"gate_{m}"]
    out["disagree_focus"] = out["focus_H"] != out[f"focus_{m}"]
    out["disagree_theme"] = out["theme_H"] != out[f"theme_{m}"]
    out["disagree_both_fields"] = out["disagree_focus"] | out["disagree_theme"]
    return out


def summary_counts(df: pd.DataFrame, model: str) -> pd.DataFrame:
    m = model
    label = "GPT-4o-mini" if m == "GPT" else "Gemini Pro"
    flagged = add_flags(df, m)
    rows = []
    for scope_name, sub in [
        ("ALL", flagged),
        ("Culicidae only", flagged[flagged["query_id"] == "mosquitoes"]),
        ("Rest (no Culicidae)", flagged[flagged["query_id"] != "mosquitoes"]),
    ]:
        n = len(sub)
        rows.append(
            {
                "model": label,
                "scope": scope_name,
                "n": n,
                "gate_disagreements": int(sub["disagree_gate"].sum()),
                "focus_disagreements": int(sub["disagree_focus"].sum()),
                "theme_disagreements": int(sub["disagree_theme"].sum()),
                "any_field_disagreement": int(sub["disagree_both_fields"].sum()),
                "human_keep_model_drop": int(
                    ((sub["gate_H"] == "keep") & (sub[f"gate_{m}"] == "drop")).sum()
                ),
                "human_drop_model_keep": int(
                    ((sub["gate_H"] == "drop") & (sub[f"gate_{m}"] == "keep")).sum()
                ),
            }
        )
    return pd.DataFrame(rows)


def top_patterns(series: pd.Series, n: int = 15) -> pd.DataFrame:
    vc = series.value_counts().head(n).reset_index()
    vc.columns = ["pattern", "count"]
    return vc


def write_sheet(ws, df: pd.DataFrame, freeze_header: bool = True) -> None:
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9E2F3")
    if freeze_header:
        ws.freeze_panes = "A2"
    # modest column widths
    widths = {
        "A": 12,
        "B": 14,
        "C": 6,
        "D": 10,
        "E": 8,
        "F": 28,
        "G": 50,
        "H": 55,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_workbook(
    joined: pd.DataFrame,
    out_xlsx: Path,
    out_dir: Path,
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    # README
    ws = wb.create_sheet("00_README")
    readme = [
        ["Human vs LLM disagreement review sheets"],
        [""],
        ["Purpose", "Help reviewers inspect coding errors by agreement category."],
        ["Human reference", "Expert labels in human_labeled_1376_20260830.csv"],
        ["Models", "GPT-4o-mini (GPT) and Gemini Pro (GEM)"],
        [""],
        ["Sheet guide"],
        ["01_Summary", "Counts of disagreements by scope and model"],
        ["GPT_1_gate", "Keep vs drop disagreements (GPT vs human)"],
        ["GPT_2_focus", "Taxon focus label disagreements (GPT vs human)"],
        ["GPT_3_theme", "Research theme disagreements (GPT vs human)"],
        ["GPT_4_both", "Focus OR theme disagreements (GPT vs human)"],
        ["GEM_*", "Same four sheets for Gemini Pro"],
        ["Culicidae_GPT", "All GPT disagreements, Culicidae only"],
        ["Culicidae_GEM", "All Gemini disagreements, Culicidae only"],
        ["Both_models_vs_H", "Both models disagree with human on focus"],
        [""],
        ["Columns", "focus_pattern / theme_pattern / gate_pattern show Human → Model"],
        ["reviewer_notes", "Empty column for your comments"],
    ]
    for row in readme:
        ws.append(row)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70

    summary_parts = []
    sheet_specs: list[tuple[str, pd.DataFrame]] = []

    for model, prefix in [("GPT", "GPT"), ("GEM", "GEM")]:
        flagged = add_flags(joined, model)
        rev = review_row(joined, model)
        summary_parts.append(summary_counts(joined, model))

        gate_mask = flagged["disagree_gate"]
        focus_mask = flagged["disagree_focus"]
        theme_mask = flagged["disagree_theme"]
        any_mask = flagged["disagree_both_fields"]

        sheet_specs.extend(
            [
                (f"{prefix}_1_gate", rev[gate_mask].sort_values(["taxon", "coding_row_id"])),
                (f"{prefix}_2_focus", rev[focus_mask].sort_values(["taxon", "coding_row_id"])),
                (f"{prefix}_3_theme", rev[theme_mask].sort_values(["taxon", "coding_row_id"])),
                (f"{prefix}_4_both", rev[any_mask].sort_values(["taxon", "coding_row_id"])),
            ]
        )

        cul_mask = (joined["query_id"] == "mosquitoes") & any_mask
        cul = rev[cul_mask].sort_values("coding_row_id")
        sheet_specs.append((f"Culicidae_{prefix}", cul))

        # CSV exports per category
        for name, sub in [
            ("gate", rev[gate_mask]),
            ("focus", rev[focus_mask]),
            ("theme", rev[theme_mask]),
            ("any_field", rev[any_mask]),
        ]:
            csv_name = f"{prefix.lower()}_{name}_disagreements.csv"
            sub.to_csv(out_dir / csv_name, index=False)

    # Both models wrong on focus vs human
    both_wrong = joined[
        (joined["focus_H"] != joined["focus_GPT"])
        & (joined["focus_H"] != joined["focus_GEM"])
    ]
    sheet_specs.append(
        (
            "Both_models_vs_H",
            review_row(both_wrong, "GPT").sort_values(["taxon", "coding_row_id"]),
        )
    )
    review_row(both_wrong, "GPT").to_csv(
        out_dir / "both_models_focus_disagree_vs_human.csv", index=False
    )

    summary = pd.concat(summary_parts, ignore_index=True)
    sheet_specs.insert(0, ("01_Summary", summary))

    # Pattern tabs for gate/focus on Culicidae GPT
    cul_gpt = joined[joined["query_id"] == "mosquitoes"]
    cul_flag = add_flags(cul_gpt, "GPT")
    cul_rev = review_row(cul_gpt, "GPT")
    for label, col in [
        ("GPT_Culi_gate_patterns", "gate_pattern"),
        ("GPT_Culi_focus_patterns", "focus_pattern"),
        ("GPT_Culi_theme_patterns", "theme_pattern"),
    ]:
        sub = cul_rev[cul_flag["disagree_gate" if "gate" in label else "disagree_focus" if "focus" in label else "disagree_theme"]]
        if len(sub):
            sheet_specs.append((label, top_patterns(sub[col], 25)))

    for sheet_name, frame in sheet_specs:
        ws = wb.create_sheet(sheet_name[:31])  # Excel sheet name limit
        if len(frame) == 0:
            ws.append(["No disagreements in this category."])
        else:
            write_sheet(ws, frame)

    wb.save(out_xlsx)

    # README markdown
    (out_dir / "README.md").write_text(
        "\n".join(
            [
                "# Disagreement review sheets",
                "",
                f"Excel workbook: `{out_xlsx.name}`",
                "",
                "## Sheets",
                "- **GPT_1_gate** / **GEM_1_gate** — keep vs drop only",
                "- **GPT_2_focus** / **GEM_2_focus** — taxon focus label",
                "- **GPT_3_theme** / **GEM_3_theme** — research theme",
                "- **GPT_4_both** / **GEM_4_both** — any focus or theme mismatch",
                "- **Culicidae_GPT** / **Culicidae_GEM** — all disagreements, Culicidae",
                "- **Both_models_vs_H** — human focus differs from both models",
                "",
                "CSV copies of main GPT/GEM category files are in this folder.",
                "",
                "## For reviewers",
                "Use `reviewer_notes` to flag whether the human or model label looks correct.",
            ]
        ),
        encoding="utf-8",
    )


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--joined", type=Path, default=DEFAULT_JOINED)
    p.add_argument("--human", type=Path, default=DEFAULT_HUMAN)
    p.add_argument("--out-dir", type=Path, default=DEFAULT_OUT)
    args = p.parse_args()

    joined = pd.read_csv(args.joined)
    human = pd.read_csv(args.human)
    df = enrich(joined, human)

    xlsx = args.out_dir / "human_disagreements_by_category.xlsx"
    build_workbook(df, xlsx, args.out_dir)
    print(f"Wrote {xlsx}")
    print(f"Wrote CSVs + README in {args.out_dir}")
    for model in ("GPT", "GEM"):
        flagged = add_flags(df, model)
        print(
            f"  {model}: gate={flagged.disagree_gate.sum()} "
            f"focus={flagged.disagree_focus.sum()} "
            f"theme={flagged.disagree_theme.sum()}"
        )


if __name__ == "__main__":
    main()
